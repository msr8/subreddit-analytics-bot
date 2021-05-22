from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import colorama
from colorama import Fore, Back, Style
import time as t
import praw
import os
colorama.init()




def handle_value(num):
    ret = num
    if num > 1000000:
        ret = round(num/1000000, 2)
        ret = str(ret) + 'M'
    elif num > 100000:
        ret = num // 1000
        ret = str(ret) + 'k'
    elif num > 10000:
        ret = round(num/1000, 1)
        ret = str(ret) + 'k'
    elif num > 1000:
        ret = round(num/1000, 2)
        ret = str(ret) + 'k'
    return ret

def sort_dict(dict):
    new_dict = {}
    new_list = sorted(dict.items(), key=lambda x: x[1])
    new_list.reverse()
    for i in new_list:
        new_dict[i[0]] = i[1]
    return new_dict

def handle_post_count(post):
    op = post.author
    op = str(op).lower()
    # Checks if op has deleted their acc
    if not op == 'none':
        # Makes a count of how much posts an user has made
        if op not in ops.keys():
            ops[op] = 1
        else:
            ops[op] += 1

def handle_word_count(post):
    for word in post.title.split(' '):
        word = word.lower()
        if len(word):
            if word not in word_count.keys():
                word_count[word] = 1
            else:
                word_count[word] += 1

def handle_upvote_count(post):
    op = post.author
    op = str(op).lower()
    # Checks if OP has deleted their acc
    if not op == 'None':
        if op not in upvotes_count.keys():
            upvotes_count[op] = post.ups
        else:
            upvotes_count[op] += post.ups

def handle_total_data(post, rank):
    temp_dict = {'op': str(post.author).lower(),
    'title': post.title,
    'ups': str(post.ups),
    'coms': str(post.num_comments),
    'created': t.asctime(t.localtime(post.created)),
    'id': post.id,
    'link': post.url}
    total_data[str(rank)] = temp_dict

def make_excel():
    wb = Workbook()
    sheet_main = wb.active
    # General Info
    sheet_main.title = 'General Info'
    sheet_main.append(['Subreddit Info'])
    sheet_main.append(['Subreddit', sub.display_name])
    sheet_main.append(['Total Members', str(sub.subscribers)])
    sheet_main.append(['18+', over18])
    sheet_main.append([])
    sheet_main.append(['Averages'])
    sheet_main.append(['Average Upvotes', str(total_ups//count)])
    sheet_main.append(['Average Comments', str(total_coms//num)])
    sheet_main.append(['Average posting time', t.asctime(t.localtime(total_time//count)) ])
    sheet_main.append([])
    sheet_main.append(['Query Info'])
    sheet_main.append(['Time posted', query_time])
    sheet_main.append(['No. of posts', str(num)])
    # Makes the cells bold
    for col in range(1,14):
        sheet_main[f'A{col}'].font = Font(bold=True)
    # Makes the cells Red
    for cell in ['A1', 'A6', 'A11']:
        sheet_main[cell].font = Font(color='EB4034', bold=True)
    # Merges the cells
    for cell_range in ['A1:B1', 'A6:B6', 'A11:B11']:
        sheet_main.merge_cells(cell_range)

    # Top Posts
    wb.create_sheet('Top Posts')
    sheet_posts = wb['Top Posts']
    sheet_posts.append(['Rank', 'OP', 'Upvotes', 'Comments', 'Time Created', 'ID', 'Link', 'Title'])
    sheet_posts.append([])
    # Writes the data
    for rank in total_data.keys():
        sheet_posts.append( [rank, total_data[rank]['op'], total_data[rank]['ups'], total_data[rank]['coms'], total_data[rank]['created'], total_data[rank]['id'], total_data[rank]['link'], total_data[rank]['title']  ] )
    # Makes the 1st row bold
    for col in range(1,9):
        sheet_posts[f'{get_column_letter(col)}1'].font = Font(bold=True)

    # Top Posters
    wb.create_sheet('Top Posters')
    sheet_posters = wb['Top Posters']
    sheet_posters.append(['Rank', 'OP', 'No. of Posts', 'Total Upvotes', 'Link'])
    sheet_posters.append([])
    poster_rank = 1
    # Writes the data
    for op in ops.keys():
        sheet_posters.append([str(poster_rank), op, str(ops[op]), str(upvotes_count[op]), f'https://reddit.com/u/{op}'])
        poster_rank += 1
    # Makes the first row bold
    for col in range(1,6):
        sheet_posters[f'{get_column_letter(col)}1'].font = Font(bold=True)
    
    # Top Words
    wb.create_sheet('Top Words')
    sheet_words = wb['Top Words']
    sheet_words.append(['Rank', 'Word', 'Occurance'])
    sheet_words.append([])
    word_rank = 1
    # Writes the data
    for word in word_count.keys():
        sheet_words.append([ str(word_rank), word, str(word_count[word]) ])
        word_rank += 1
    # Makes the first row bold
    for col in range(1,4):
        sheet_words[f'{get_column_letter(col)}1'].font = Font(bold=True)



    xl_file_path = os.path.join( os.path.dirname(__file__),'Output', f'{sub.display_name}.xlsx' )
    if not os.path.exists( os.path.dirname(xl_file_path) ):
        os.makedirs( os.path.dirname(xl_file_path) )
    wb.save(xl_file_path)
    print(f'Saved the data at {Fore.GREEN}{xl_file_path}{Fore.RESET}\n')








start = t.perf_counter()
query_time = t.asctime()

config_file_path = os.path.join( os.path.dirname(__file__),'config.txt' )

# Gets the client id and secret
with open(config_file_path, 'r') as config:
    _, client_id, _, client_secret, _, is_ask_sub, _, sub, _ = config.read().split('"')

if is_ask_sub == 'False':
    is_ask_sub = False
else:
    is_ask_sub = True

reddit = praw.Reddit(client_id = client_id,
client_secret = client_secret,
user_agent = 'wtv')

if is_ask_sub:
    sub = input('\nEnter the subreddit you want to run the query on:\n')

if sub.startswith('r/'):
    sub = sub[2:]
elif sub.startswith('/'):
    sub = sub[1:]
sub = reddit.subreddit(sub)

num = 1000
total_ups = 0
total_coms = 0
total_time = 0
if sub.over18:
    over18 = 'Yes'
else:
    over18 = 'No'

ops = {}
word_count = {}
upvotes_count = {}
total_data = {}
count = 0
rank_count = 1



# Goes through the top 1000 posts
for post in sub.top(limit=num):
    total_ups += post.ups
    total_coms += post.num_comments
    total_time += post.created
    op = post.author
    op = str(op).lower()
    handle_post_count(post)
    handle_word_count(post)
    handle_upvote_count(post)
    handle_total_data(post, rank_count)
    count += 1
    rank_count += 1

# Sorts the dictionaries
ops = sort_dict(ops)
word_count = sort_dict(word_count)

# Sub info
print(f'\n{Fore.YELLOW}============================== SUBREDDIT INFO ==============================\n{Fore.RESET}')
print(f'{Fore.CYAN}Subreddit:{Fore.RESET} r/{sub.display_name}')
print(f'{Fore.CYAN}Total Members:{Fore.RESET} {handle_value(sub.subscribers)}')
print(f'{Fore.CYAN}18+ :{Fore.RESET} {over18}')
print(f'{Fore.CYAN}Time created:{Fore.RESET} {t.asctime(t.localtime(sub.created))}')
print(f'{Fore.CYAN}Posts gone through:{Fore.RESET} {count}')

# Top post info
for top_post in sub.top(limit=1):
    print(f'\n{Fore.YELLOW}============================== TOP POST ==============================\n{Fore.RESET}')
    print(f'{Fore.CYAN}Title:{Fore.RESET} {top_post.title}')
    print(f'{Fore.CYAN}OP:{Fore.RESET} u/{ ( str(top_post.author) ).lower() }')
    print(f'{Fore.CYAN}Upvotes:{Fore.RESET} {handle_value(top_post.ups)}')
    print(f'{Fore.CYAN}Comments:{Fore.RESET} {top_post.num_comments}')
    print(f'{Fore.CYAN}Time posted:{Fore.RESET} {t.asctime(t.localtime(top_post.created))}')

# Averages
print(f'\n{Fore.YELLOW}============================== AVERAGES ==============================\n{Fore.RESET}')
print(f'{Fore.CYAN}Average upvotes of the top {count} posts:{Fore.RESET} {handle_value(total_ups//count)}')
print(f'{Fore.CYAN}Average comments of the top {count} posts:{Fore.RESET} {handle_value(total_coms//count)}')
print(f'{Fore.CYAN}Average time of posting:{Fore.RESET} {t.asctime(t.localtime(total_time//count))}')

# Top OPs
poster_count = 1
print(f'\n{Fore.YELLOW}============================== TOP POSTERS =============================={Fore.RESET}')
print(f'\n{Fore.RED}Rank{Fore.RESET} ) {Fore.RED}OP{Fore.RESET} : {Fore.RED}Number of posts{Fore.RESET} : {Fore.RED}Total Upvotes Gained{Fore.RESET}\n')
for key in list(ops.keys())[0:10]:
    print(f'{poster_count}) u/{key} : {ops[key]} : {handle_value(upvotes_count[key])}')
    poster_count += 1
print(f'\n.........and {len(ops) - 10} more')

# Top Words
word_count_int = 1
print(f'\n{Fore.YELLOW}============================== TOP WORDS USED =============================={Fore.RESET}')
print(f'\n{Fore.RED}Rank{Fore.RESET} ) {Fore.RED}Word{Fore.RESET} : {Fore.RED}Times the word has been used{Fore.RESET}\n')
for key in list(word_count.keys())[0:10]:
    print(f'{word_count_int}) {key} : {word_count[key]}')
    word_count_int += 1
print(f'\n.........and {len(word_count) - 10} more')





print(f'\n\nTime taken to perform the queries: {round(t.perf_counter() - start, 2)} seconds')
print('\nSaving the data\n')

make_excel()


